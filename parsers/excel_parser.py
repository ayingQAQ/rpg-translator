"""
Excel Parser - Parse and reconstruct Excel files
"""

from typing import List, Tuple, Any, Dict, Optional
from .base_parser import BaseParser, TextSegment


class ExcelParser(BaseParser):
    """Parser for Excel files (.xlsx, .xls)."""
    
    def __init__(
        self,
        file_path: str,
        sheet_index: int = 0,
        text_columns: List[str] = None,
        header_row: bool = True,
        **kwargs
    ):
        """
        Initialize Excel parser.
        
        Args:
            file_path: Path to Excel file
            sheet_index: Sheet index to process (0-based)
            text_columns: Column letters to translate (e.g., ['A', 'C'])
            header_row: Whether first row is header
            **kwargs: Additional options
        """
        super().__init__(file_path, **kwargs)
        self.sheet_index = sheet_index
        self.text_columns = text_columns
        self.header_row = header_row
        self.workbook = None
        self.sheet = None
        self.detected_columns = []
    
    def _col_letter_to_index(self, letter: str) -> int:
        """Convert column letter to 0-based index (A=0, B=1, ...)."""
        result = 0
        for char in letter.upper():
            result = result * 26 + (ord(char) - ord('A') + 1)
        return result - 1
    
    def _index_to_col_letter(self, index: int) -> str:
        """Convert 0-based index to column letter."""
        result = ""
        index += 1  # Convert to 1-based
        while index > 0:
            index, remainder = divmod(index - 1, 26)
            result = chr(ord('A') + remainder) + result
        return result
    
    def _detect_text_columns(self) -> List[int]:
        """
        Detect which columns contain translatable text.
        
        Similar logic to CSV parser.
        """
        max_row = min(self.sheet.max_row, 100)  # Sample first 100 rows
        max_col = self.sheet.max_column
        
        column_scores = []
        
        for col_idx in range(1, max_col + 1):
            texts = []
            for row_idx in range(2 if self.header_row else 1, max_row + 1):
                cell = self.sheet.cell(row=row_idx, column=col_idx)
                if cell.value and isinstance(cell.value, str):
                    texts.append(cell.value)
            
            if not texts:
                column_scores.append((col_idx - 1, 0))
                continue
            
            # Calculate score
            avg_len = sum(len(t) for t in texts) / len(texts)
            numeric_ratio = sum(1 for t in texts if t.strip().isdigit()) / len(texts)
            score = avg_len * (1 - numeric_ratio)
            
            column_scores.append((col_idx - 1, score))
        
        # Select columns with score above threshold
        text_cols = [
            idx for idx, score in column_scores
            if score > 3
        ]
        
        return text_cols if text_cols else list(range(max_col))
    
    def parse(self) -> List[TextSegment]:
        """Parse Excel file and extract text from specified columns."""
        
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel support.\n"
                "Install it with: pip install openpyxl"
            )
        
        # Load workbook
        self.workbook = load_workbook(self.file_path)
        
        # Get sheet
        if self.sheet_index >= len(self.workbook.worksheets):
            raise ValueError(f"Sheet index {self.sheet_index} out of range")
        
        self.sheet = self.workbook.worksheets[self.sheet_index]
        
        # Detect text columns if not specified
        if self.text_columns is None:
            self.detected_columns = self._detect_text_columns()
        else:
            self.detected_columns = [
                self._col_letter_to_index(col) for col in self.text_columns
            ]
        
        # Extract text segments
        segments = []
        start_row = 2 if self.header_row else 1
        
        for row_idx in range(start_row, self.sheet.max_row + 1):
            for col_idx in self.detected_columns:
                cell = self.sheet.cell(row=row_idx, column=col_idx + 1)
                text = cell.value
                
                if text and isinstance(text, str) and not self.should_skip_text(text):
                    col_letter = self._index_to_col_letter(col_idx)
                    segments.append(TextSegment(
                        text=text,
                        location=f"{col_letter}{row_idx}",
                        metadata={
                            'row': row_idx,
                            'column': col_idx + 1,
                            'column_letter': col_letter,
                            'header': self.sheet.cell(row=1, column=col_idx + 1).value if self.header_row else None
                        }
                    ))
        
        return segments
    
    def reconstruct(self, translated_segments: List[Tuple[str, str]]) -> Any:
        """
        Reconstruct Excel with translations.
        
        Args:
            translated_segments: List of (original, translated) tuples
        
        Returns:
            Modified workbook
        """
        if self.workbook is None:
            raise ValueError("No original data. Call parse() first.")
        
        # Create translation mapping
        translation_map = {orig: trans for orig, trans in translated_segments}
        
        # Apply translations
        start_row = 2 if self.header_row else 1
        
        for row_idx in range(start_row, self.sheet.max_row + 1):
            for col_idx in self.detected_columns:
                cell = self.sheet.cell(row=row_idx, column=col_idx + 1)
                if cell.value and str(cell.value) in translation_map:
                    cell.value = translation_map[str(cell.value)]
        
        return self.workbook
    
    def save(self, workbook: Any, output_path: str) -> None:
        """Save workbook to file."""
        import os
        os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
        workbook.save(output_path)
